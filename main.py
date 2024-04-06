from django.http import HttpResponse
from openpyxl import Workbook

def export_to_excel(request):
    # Create a new Excel workbook
    wb = Workbook()

    # Create worksheets for users and tasks
    user_ws = wb.create_sheet(title="Users")
    task_ws = wb.create_sheet(title="Tasks")

    # Write headers for users and tasks
    user_ws.append(['ID', 'Name', 'Email', 'Mobile'])
    task_ws.append(['User', 'Task Detail', 'Task Type'])

    # Write user data
    for user in User.objects.all():
        user_ws.append([user.id, user.name, user.email, user.mobile])

    # Write task data
    for task in Task.objects.all():
        task_ws.append([task.user.name, task.task_detail, task.task_type])

    # Create a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Set file name
    response['Content-Disposition'] = 'attachment; filename=myapp_data.xlsx'

    # Save the workbook content to the response
    wb.save(response)

    return response
